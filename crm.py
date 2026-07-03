#!/usr/bin/env python3
"""Personal CRM — who to contact today, templates, logging."""

import sqlite3
import argparse
import re
import csv
from datetime import datetime, date, timedelta
from pathlib import Path

from config import DB, CRM_XLSX, MEETINGS_XLSX, LINKEDIN_CONNECTIONS, LINKEDIN_MESSAGES

TODAY = date.today()


# ─── DB setup ────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn


def init_db(conn):
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS contacts (
        id          INTEGER PRIMARY KEY,
        name        TEXT NOT NULL,
        title       TEXT,
        company     TEXT,
        priority    TEXT,   -- HOT / Stanford / Warm / Gartner / New / Cold / Investor
        status      TEXT,
        last_contact_raw TEXT,
        last_contact_days INTEGER,  -- days ago (computed)
        channel     TEXT,
        why_icp     TEXT,
        source      TEXT,
        next_action TEXT,
        email       TEXT,
        linkedin    TEXT,
        notes       TEXT
    );
    CREATE TABLE IF NOT EXISTS meetings (
        id      INTEGER PRIMARY KEY,
        date    TEXT,
        name    TEXT,
        email   TEXT,
        title   TEXT
    );
    CREATE TABLE IF NOT EXISTS log (
        id         INTEGER PRIMARY KEY,
        ts         TEXT DEFAULT (datetime('now','localtime')),
        contact_id INTEGER,
        action     TEXT,   -- email_sent / meeting_confirmed / call_done / note
        note       TEXT,
        FOREIGN KEY (contact_id) REFERENCES contacts(id)
    );
    CREATE TABLE IF NOT EXISTS linkedin (
        id              INTEGER PRIMARY KEY,
        name            TEXT NOT NULL,
        first_name      TEXT,
        last_name       TEXT,
        company         TEXT,
        title           TEXT,
        email           TEXT,
        connected_on    TEXT,
        last_msg_date   TEXT,
        last_msg_days   INTEGER,
        last_msg_snippet TEXT,
        last_direction  TEXT   -- sent / received
    );
    """)
    conn.commit()


# ─── Import ───────────────────────────────────────────────────────────────────

PRIORITY_ORDER = {
    "01-HOT": ("HOT", 1),
    "02-Stanford": ("Stanford", 2),
    "03-Gartner": ("Gartner", 3),
    "04-Warm": ("Warm", 4),
    "05-New Leads": ("New", 5),
    "06-Cold Enterprise": ("Cold", 6),
    "07-Investor": ("Investor", 7),
}

STATUS_URGENCY = {
    "damage control": 100,
    "missed time slot": 90,
    "awaits response": 80,
    "waiting for your response": 75,
    "need re-engagement": 40,
    "not met yet": 30,
    "fresh warm": 50,
    "asked about raise": 35,
    "strong relationship": 20,
    "cold but research-ready": 15,
    "not engaged yet": 10,
}


def parse_days(raw) -> "int | None":
    """Convert Excel value like '60д назад', '~40д назад', 'Never' → int days."""
    if not raw:
        return None
    raw = str(raw).strip().lower()
    if "never" in raw or raw == "не":
        return 999
    m = re.search(r"~?(\d+)\s*д", raw)  # Russian Excel format: "Nд назад"
    if m:
        return int(m.group(1))
    # "Me2We not done", "Never deep" etc.
    if any(w in raw for w in ["never", "not done", "deep"]):
        return 999
    return None


def load_excel(conn):
    import openpyxl
    conn.execute("DELETE FROM contacts")
    conn.execute("DELETE FROM meetings")

    # ── CRM ──
    wb = openpyxl.load_workbook(CRM_XLSX)
    seen_names = set()

    for sheet_name, (priority, _) in PRIORITY_ORDER.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or not row[0]:
                continue
            name, title, company, status, last_raw, channel, why_icp, source, next_action = (
                (row[i] if i < len(row) else None) for i in range(9)
            )
            if not name or name in seen_names:
                continue
            seen_names.add(name)
            days = parse_days(last_raw)
            conn.execute(
                """INSERT INTO contacts
                   (name, title, company, priority, status, last_contact_raw,
                    last_contact_days, channel, why_icp, source, next_action)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (name, title, company, priority, status, str(last_raw) if last_raw else None,
                 days, channel, why_icp, source, next_action),
            )

    # ── Meetings ──
    wb2 = openpyxl.load_workbook(MEETINGS_XLSX)
    ws2 = wb2.active
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        dt, name, email, title = (row[i] if i < len(row) else None for i in range(4))
        if dt and name:
            conn.execute(
                "INSERT INTO meetings (date, name, email, title) VALUES (?,?,?,?)",
                (str(dt)[:10], name, email, title),
            )
    conn.commit()
    total = conn.execute("SELECT COUNT(*) FROM contacts").fetchone()[0]
    meetings = conn.execute("SELECT COUNT(*) FROM meetings").fetchone()[0]
    print(f"✓ Loaded {total} contacts, {meetings} meetings")


def load_linkedin(conn):
    today = date.today()

    # ── Messages: find last msg date per person ──
    last_msg = {}
    with open(LINKEDIN_MESSAGES, encoding="utf-8") as f:
        for m in csv.DictReader(f):
            msg_date = m["DATE"][:10]
            sender = m["FROM"].strip()
            recipient = m["TO"].strip()
            content = m["CONTENT"][:100].replace("\n", " ")
            direction = "sent" if sender == "Mariia Potupchik" else "received"
            other = recipient if sender == "Mariia Potupchik" else sender
            if other and other != "Mariia Potupchik":
                if other not in last_msg or msg_date > last_msg[other][0]:
                    last_msg[other] = (msg_date, content, direction)

    def days_ago(date_str):
        try:
            return (today - datetime.strptime(date_str, "%Y-%m-%d").date()).days
        except:
            return None

    # ── Connections ──
    conn.execute("DELETE FROM linkedin")
    with open(LINKEDIN_CONNECTIONS, encoding="utf-8") as f:
        lines = f.readlines()
    start = next(i for i, l in enumerate(lines) if l.startswith("First Name"))
    inserted = 0
    for row in csv.DictReader(lines[start:]):
        first = row["First Name"].strip()
        last = row["Last Name"].strip()
        name = f"{first} {last}".strip()
        if not name:
            continue
        msg_info = last_msg.get(name)
        conn.execute(
            """INSERT INTO linkedin
               (name, first_name, last_name, company, title, email,
                connected_on, last_msg_date, last_msg_days, last_msg_snippet, last_direction)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (name, first, last,
             row.get("Company", ""), row.get("Position", ""), row.get("Email Address", ""),
             row.get("Connected On", ""),
             msg_info[0] if msg_info else None,
             days_ago(msg_info[0]) if msg_info else None,
             msg_info[1] if msg_info else None,
             msg_info[2] if msg_info else None)
        )
        inserted += 1

    # ── Sync last_contact_days for existing CRM contacts ──
    crm_contacts = conn.execute("SELECT id, name FROM contacts").fetchall()
    updated = 0
    for c in crm_contacts:
        # fuzzy match: first + last name overlap
        first_word = c["name"].split()[0].lower()
        li = conn.execute(
            """SELECT last_msg_date, last_msg_days FROM linkedin
               WHERE lower(name) LIKE ? AND last_msg_date IS NOT NULL
               ORDER BY last_msg_date DESC LIMIT 1""",
            (f"%{first_word}%",)
        ).fetchone()
        if li and li["last_msg_days"] is not None:
            # only update if LinkedIn shows more recent contact
            existing = conn.execute(
                "SELECT last_contact_days FROM contacts WHERE id=?", (c["id"],)
            ).fetchone()
            if existing["last_contact_days"] is None or li["last_msg_days"] < existing["last_contact_days"]:
                conn.execute(
                    "UPDATE contacts SET last_contact_days=?, last_contact_raw=? WHERE id=?",
                    (li["last_msg_days"], f"{li['last_msg_days']}d ago (LinkedIn)", c["id"])
                )
                updated += 1

    conn.commit()
    total_li = conn.execute("SELECT COUNT(*) FROM linkedin").fetchone()[0]
    with_msgs = conn.execute("SELECT COUNT(*) FROM linkedin WHERE last_msg_date IS NOT NULL").fetchone()[0]
    print(f"✓ LinkedIn: {total_li} connections, {with_msgs} with messages, {updated} CRM contacts updated")


def search_linkedin(conn, query: str, limit=15):
    rows = conn.execute(
        """SELECT * FROM linkedin
           WHERE name LIKE ? OR company LIKE ? OR title LIKE ?
           ORDER BY last_msg_date DESC NULLS LAST
           LIMIT ?""",
        (f"%{query}%", f"%{query}%", f"%{query}%", limit)
    ).fetchall()
    if not rows:
        print("Nothing found.")
        return
    print(f"\n{'─'*65}")
    for r in rows:
        days = f"{r['last_msg_days']}d ago" if r['last_msg_days'] is not None else "no messages"
        direction = f"({r['last_direction']})" if r['last_direction'] else ""
        print(f"  {r['name']:<28} {r['company'] or '—':<22} {days} {direction}")
        if r["last_msg_snippet"]:
            print(f"    └ \"{r['last_msg_snippet'][:70]}\"")
    print()


# ─── Scoring ──────────────────────────────────────────────────────────────────

PRIORITY_SCORE = {"HOT": 400, "Stanford": 200, "Gartner": 150,
                  "Warm": 120, "New": 100, "Investor": 80, "Cold": 50}


def contact_score(c) -> int:
    score = PRIORITY_SCORE.get(c["priority"], 0)

    # status urgency
    status_low = (c["status"] or "").lower()
    for keyword, pts in STATUS_URGENCY.items():
        if keyword in status_low:
            score += pts
            break

    # days since last contact (more days = higher score, capped at 300)
    days = c["last_contact_days"]
    if days is not None:
        score += min(days * 3, 300)

    return score


# ─── Display ──────────────────────────────────────────────────────────────────

TEMPLATES = {
    "damage control": """\
Subject: [Apology + new slots]

Hi {name},

Sincere apologies — I dropped the ball on our timing. Been dealing with [illness/crunch].

Would any of these work for a quick call?
• [Option 1]
• [Option 2]
• [Option 3]

Totally understand if the window has passed — just let me know.

Best,
Mariia""",

    "awaits response": """\
Subject: Following up — [topic]

Hi {name},

Just circling back on my last note. I know inboxes get buried.

[1 sentence on why now is relevant — new traction / milestone / mutual connection]

Happy to keep it to 20 min. Does [date] work?

Best,
Mariia""",

    "missed time slot": """\
Subject: Apologies + rescheduling

Hi {name},

I'm so sorry I missed our scheduled time — completely on me.

Are you open to reconnecting? I have availability:
• [Slot 1]
• [Slot 2]

Best,
Mariia""",

    "default": """\
Subject: [Personalized hook]

Hi {name},

[Opening tied to their work / recent post / shared context]

I'm building YourSkills — [1-line pitch relevant to them].

Would love 20 min to [specific ask]. Does [date] work?

Best,
Mariia""",
}


def pick_template(status) -> str:
    if not status:
        return TEMPLATES["default"]
    s = status.lower()
    for key in TEMPLATES:
        if key in s:
            return TEMPLATES[key]
    return TEMPLATES["default"]


def fmt_days(days) -> str:
    if days is None:
        return "?"
    if days >= 999:
        return "never"
    return f"{days}d ago"


def show_today(conn, limit=10):
    rows = conn.execute("SELECT * FROM contacts WHERE skip=0 OR skip IS NULL").fetchall()
    ranked = sorted(rows, key=contact_score, reverse=True)[:limit]

    print(f"\n{'═'*60}")
    print(f"  REACH OUT TODAY — {TODAY.strftime('%d %b %Y')}")
    print(f"{'═'*60}")

    for i, c in enumerate(ranked, 1):
        score = contact_score(c)
        priority_emoji = {"HOT": "🔥", "Stanford": "🎓", "Gartner": "🏢",
                          "Warm": "🤝", "New": "🆕", "Investor": "💰", "Cold": "❄️"}.get(c["priority"], "•")
        print(f"\n{i}. [{priority_emoji} {c['priority']}] {c['name']} · {c['company'] or '—'}")
        print(f"   {c['title'] or ''}")
        print(f"   Status: {c['status'] or '—'}")
        print(f"   Last contact: {fmt_days(c['last_contact_days'])} | Channel: {c['channel'] or '—'}")
        print(f"   Score: {score}")
        if c["next_action"]:
            action = c["next_action"][:120] + "…" if len(c["next_action"] or "") > 120 else c["next_action"]
            print(f"   → {action}")

    print(f"\n{'─'*60}\n")


def show_contact(conn, name_query: str):
    row = conn.execute(
        "SELECT * FROM contacts WHERE name LIKE ? LIMIT 1",
        (f"%{name_query}%",)
    ).fetchone()

    if not row:
        print(f"Contact '{name_query}' not found")
        return

    print(f"\n{'═'*60}")
    print(f"  {row['name']}")
    print(f"{'═'*60}")
    print(f"  Company:  {row['company']} · {row['title']}")
    print(f"  Priority: {row['priority']}")
    print(f"  Status:   {row['status']}")
    print(f"  Contact:  {fmt_days(row['last_contact_days'])} · {row['channel']}")
    print(f"  ICP:      {row['why_icp']}")
    print(f"  Source:   {row['source']}")
    print(f"\n  Next action:\n  {row['next_action']}")

    meetings = conn.execute(
        "SELECT * FROM meetings WHERE name LIKE ? ORDER BY date DESC LIMIT 5",
        (f"%{name_query.split()[0]}%",)
    ).fetchall()
    if meetings:
        print(f"\n  Meetings:")
        for m in meetings:
            print(f"    {m['date']} — {m['title']}")

    logs = conn.execute(
        "SELECT * FROM log WHERE contact_id = ? ORDER BY ts DESC LIMIT 5",
        (row["id"],)
    ).fetchall()
    if logs:
        print(f"\n  Action history:")
        for l in logs:
            print(f"    {l['ts'][:16]} [{l['action']}] {l['note'] or ''}")

    template = pick_template(row["status"])
    print(f"\n  {'─'*50}")
    print("  TEMPLATE:")
    print("  " + template.replace("\n", "\n  ").replace("{name}", row["name"].split()[0]))
    print()


def skip_contact(conn, name_query: str, undo: bool = False):
    row = conn.execute(
        "SELECT id, name FROM contacts WHERE name LIKE ? LIMIT 1",
        (f"%{name_query}%",)
    ).fetchone()
    if not row:
        print(f"Contact '{name_query}' not found")
        return
    conn.execute("UPDATE contacts SET skip=? WHERE id=?", (0 if undo else 1, row["id"]))
    conn.commit()
    action = "Restored" if undo else "Skipped (vendor)"
    print(f"✓ {action}: {row['name']}")


def log_action(conn, name_query: str, action: str, note: str):
    row = conn.execute(
        "SELECT id, name FROM contacts WHERE name LIKE ? LIMIT 1",
        (f"%{name_query}%",)
    ).fetchone()
    if not row:
        print(f"Contact '{name_query}' not found")
        return
    conn.execute(
        "INSERT INTO log (contact_id, action, note) VALUES (?,?,?)",
        (row["id"], action, note)
    )
    conn.execute(
        "UPDATE contacts SET last_contact_days=0, last_contact_raw='today' WHERE id=?",
        (row["id"],)
    )
    conn.commit()
    print(f"✓ Logged: {row['name']} → [{action}] {note}")


def show_log(conn, limit=20):
    rows = conn.execute(
        """SELECT l.ts, c.name, l.action, l.note
           FROM log l JOIN contacts c ON l.contact_id=c.id
           ORDER BY l.ts DESC LIMIT ?""",
        (limit,)
    ).fetchall()
    if not rows:
        print("Log is empty.")
        return
    print(f"\n{'─'*60}")
    for r in rows:
        print(f"  {r['ts'][:16]}  {r['name']:<25} [{r['action']}] {r['note'] or ''}")
    print()


def search_contacts(conn, query: str):
    rows = conn.execute(
        """SELECT * FROM contacts
           WHERE name LIKE ? OR company LIKE ? OR title LIKE ?
           ORDER BY last_contact_days DESC""",
        (f"%{query}%", f"%{query}%", f"%{query}%")
    ).fetchall()
    if not rows:
        print("Nothing found.")
        return
    for c in rows:
        print(f"  [{c['priority']}] {c['name']} · {c['company']} · {c['status']}")


# ─── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Personal CRM")
    sub = parser.add_subparsers(dest="cmd")

    sub.add_parser("import", help="Load Excel → SQLite")
    sub.add_parser("import-linkedin", help="Load LinkedIn connections + messages")

    p_today = sub.add_parser("today", help="Who to reach out to today")
    p_today.add_argument("-n", type=int, default=10, help="Top N contacts")

    p_show = sub.add_parser("show", help="Show contact card + template")
    p_show.add_argument("name", nargs="+")

    p_log = sub.add_parser("log", help="Log an action")
    p_log.add_argument("name", nargs="+")
    p_log.add_argument("-a", "--action", default="email_sent",
                       choices=["email_sent", "meeting_confirmed", "call_done", "replied", "note"])
    p_log.add_argument("-m", "--message", default="")

    sub.add_parser("history", help="Show action log")

    p_skip = sub.add_parser("skip", help="Mark contact as vendor (hide from today)")
    p_skip.add_argument("name", nargs="+")
    p_skip.add_argument("--undo", action="store_true", help="Restore skipped contact")

    p_search = sub.add_parser("search", help="Search CRM contacts")
    p_search.add_argument("query", nargs="+")

    p_li = sub.add_parser("linkedin", help="Search LinkedIn connections")
    p_li.add_argument("query", nargs="+")

    args = parser.parse_args()

    conn = get_db()
    init_db(conn)

    if args.cmd == "import":
        load_excel(conn)

    elif args.cmd == "import-linkedin":
        load_linkedin(conn)

    elif args.cmd == "today":
        show_today(conn, args.n)

    elif args.cmd == "show":
        show_contact(conn, " ".join(args.name))

    elif args.cmd == "log":
        log_action(conn, " ".join(args.name), args.action, args.message)

    elif args.cmd == "history":
        show_log(conn)

    elif args.cmd == "skip":
        skip_contact(conn, " ".join(args.name), undo=args.undo)

    elif args.cmd == "search":
        search_contacts(conn, " ".join(args.query))

    elif args.cmd == "linkedin":
        search_linkedin(conn, " ".join(args.query))

    else:
        parser.print_help()

    conn.close()


if __name__ == "__main__":
    main()
